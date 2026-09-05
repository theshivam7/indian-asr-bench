"""Build a non-blind human review sheet of Svarah clips that are hard for
several strong models at once.

Same protocol as analysis/tie_validation/build_sample.py: shows the annotator
every model's hypothesis so they can judge, per clip, whether the high WER
comes from the audio, the reference transcript, or the models themselves.
Clips are selected by requiring several of the strongest available models to
agree a clip is hard, which filters out model-specific failure modes.

Review columns, all reviewer-facing (see write_xlsx for the exact layout):
    reference_check     is the reference transcript itself correct
    corrected_reference  free text, fill in only if reference_check != Correct
    hyp_<model>_check   per-model correctness (a clip is rarely uniformly
                         right or wrong across all 5 models, so this is
                         tracked per model rather than one aggregate column)
    error_type          one cell, but a clip can have more than one cause:
                         pick from the dropdown for a single cause, or type
                         two or more options separated by commas
    reviewer_decision   final one-word verdict for the row
    reviewer_notes      free text, optional

Writes two files:
    review_sheet.csv   source of truth, plain CSV
    review_sheet.xlsx  same data, reviewer-facing, with the review columns
                        constrained to dropdown lists (data validation)

Usage:
    uv run --with openpyxl python3 analysis/svarah_validation/build_sample.py
"""

import os
import sys
import csv
import random

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from utils.io_helpers import stage2_dir

DATASET = "svarah"
MODE = "transcript_clean"
REQUIRED_MODELS = ["large", "parakeet", "parakeet_ctc", "qwen3"]
BONUS_MODEL = "medium"
ALL_MODELS = REQUIRED_MODELS + [BONUS_MODEL]
WER_THRESHOLD = 40.0  # percent
MIN_MODELS_FLAGGED = 3  # of len(REQUIRED_MODELS)

# Svarah's clips are far shorter than TIE's lecture segments, so the same WER
# threshold selects a different kind of clip here. 232 of the 499 clips it flags
# have a one-word reference, where WER can only be 0% or 100%: one wrong word
# clears the 40% bar automatically. That is the metric being coarse, not the clip
# being hard. Controlling for reference length, sub-2s clips are no harder than
# longer ones (3-5 word refs: 6.2% flagged under 2s vs 7.6% over). So the sample
# is filtered on reference length rather than duration, which is only a proxy for
# it, and a leaky one: a duration cut would drop 30 valid fast-speech clips and
# keep 11 two-word ones.
MIN_REF_WORDS = 3

# 209 clips survive that filter, still too many to hand-review. Sample within
# duration bands so the reviewed set spans the corpus instead of piling up in the
# shortest band. Counts are ~29% of each band, giving 60 clips, comparable to the
# 49 reviewed for TIE. Bands are (low, high] in seconds; None means no upper bound.
SAMPLE_PER_BAND = {
    (0, 2): 8,
    (2, 4): 16,
    (4, 6): 11,
    (6, 9): 11,
    (9, 14): 10,
    (14, None): 4,
}
SAMPLE_SEED = 42

HERE = os.path.dirname(__file__)
CSV_PATH = os.path.join(HERE, "review_sheet.csv")
XLSX_PATH = os.path.join(HERE, "review_sheet.xlsx")
SAMPLE_CSV_PATH = os.path.join(HERE, "review_sample.csv")
SAMPLE_XLSX_PATH = os.path.join(HERE, "review_sample.xlsx")

# Dropdown options. Kept short and closed-set on purpose: a fixed vocabulary
# is what makes the filled sheet directly tabulable afterward, instead of
# free text that needs re-normalizing before it can be analyzed.
REFERENCE_CHECK_OPTIONS = ["Correct", "Partially correct", "Incorrect"]
HYP_CHECK_OPTIONS = ["Correct", "Partially correct", "Incorrect"]
REVIEWER_DECISION_OPTIONS = [
    "Genuine model error",
    "Reference error",
    "Audio artifact",
    "Not a real error",
    "Unsure",
]

# error_type is one column but supports more than one cause per clip: the
# dropdown offers these as single picks, and typing two or more separated by
# commas (e.g. "Noise, Technical vocabulary") is accepted, just flagged with
# a soft warning instead of being blocked, since Excel/xlsx data validation
# has no native multi-select-into-one-cell control.
ERROR_TYPE_OPTIONS = [
    "Noise / audio quality",
    "Speed (fast or slow / unclear)",
    "Accent / pronunciation",
    "Technical vocabulary (jargon, numbers, names)",
    "Disfluency (fillers, repetitions, false starts)",
    "Code-switching (non-English words)",
    "Reference error",
    "Misalignment (wrong clip boundary)",
    "Other",
]


def load_model(model: str) -> dict:
    """Return {sample_id: row_dict} from that model's WER csv."""
    path = os.path.join(stage2_dir(DATASET), MODE, f"wer_{model}_{MODE}.csv")
    rows = {}
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows[str(row["ID"])] = row
    return rows


def stratified_sample(rows: list) -> list:
    """Pick SAMPLE_PER_BAND clips per duration band from the reference-filtered pool.

    Seeded, so the same pool always yields the same sheet. A band holding fewer
    clips than its quota contributes all of them rather than failing.
    """
    eligible = [r for r in rows if r["ref_words"] >= MIN_REF_WORDS]
    rng = random.Random(SAMPLE_SEED)
    picked = []
    for (lo, hi), quota in SAMPLE_PER_BAND.items():
        band = [
            r for r in eligible
            if r["duration_seconds"] not in ("", None)
            and lo <= float(r["duration_seconds"]) < (hi if hi is not None else float("inf"))
        ]
        take = sorted(band, key=lambda r: r["sample_id"])
        picked.extend(take if len(take) <= quota else rng.sample(take, quota))
        print(f"    {lo:>2}-{str(hi) + 's' if hi else 'max':<5} pool {len(band):>3}  sampled {min(len(band), quota):>3}")
    picked.sort(key=lambda r: (float(r["duration_seconds"]), r["sample_id"]))
    return picked


def main() -> None:
    tables = {m: load_model(m) for m in ALL_MODELS}
    base_ids = list(tables[REQUIRED_MODELS[0]].keys())

    flagged_counts = {m: 0 for m in ALL_MODELS}
    out_rows = []

    for sid in base_ids:
        base_row = tables[REQUIRED_MODELS[0]][sid]
        wers = {}
        hyps = {}
        for m in ALL_MODELS:
            r = tables[m].get(sid)
            if r is None:
                continue
            w = float(r["wer"]) * 100
            wers[m] = w
            hyps[m] = r["hypothesis_raw"]
            if w > WER_THRESHOLD:
                flagged_counts[m] += 1

        n_models_flagged = sum(
            1 for m in REQUIRED_MODELS if m in wers and wers[m] > WER_THRESHOLD
        )
        if n_models_flagged < MIN_MODELS_FLAGGED:
            continue

        avg_wer = round(sum(wers[m] for m in wers) / len(wers), 2) if wers else ""

        row = {
            "sample_id": sid,
            "audio_filename": f"{sid}.wav",
            "reference": base_row["reference_raw"],
            **{f"hyp_{m}": hyps.get(m, "") for m in ALL_MODELS},
            **{f"wer_{m}": round(wers[m], 2) if m in wers else "" for m in ALL_MODELS},
            "avg_wer": avg_wer,
            "n_models_flagged": n_models_flagged,
            # Svarah has no Native_Region column (that is TIE's field); its
            # closest per-speaker metadata is Native_Language.
            "native_language": base_row.get("Native_Language", ""),
            "duration_seconds": base_row.get("Speech_Duration_seconds", ""),
            # Counted on the normalized reference, the text WER is scored against,
            # not the raw one: normalization splits possessives and drops
            # punctuation-only tokens, so the two counts disagree on ~20 clips.
            "ref_words": len(base_row["reference"].split()),
            "reference_check": "",
            "corrected_reference": "",
            **{f"hyp_{m}_check": "" for m in ALL_MODELS},
            "error_type": "",
            "reviewer_decision": "",
            "reviewer_notes": "",
        }
        out_rows.append(row)

    out_rows.sort(key=lambda r: r["n_models_flagged"], reverse=True)
    for i, row in enumerate(out_rows, start=1):
        row["sr_no"] = i

    fieldnames = (
        ["sr_no", "sample_id", "audio_filename", "reference"]
        + [f"hyp_{m}" for m in ALL_MODELS]
        + [f"wer_{m}" for m in ALL_MODELS]
        + ["avg_wer", "n_models_flagged", "native_language", "duration_seconds",
           "ref_words", "reference_check", "corrected_reference"]
        + [f"hyp_{m}_check" for m in ALL_MODELS]
        + ["error_type", "reviewer_decision", "reviewer_notes"]
    )

    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    write_xlsx(fieldnames, out_rows)

    print(f"[sample] reference >= {MIN_REF_WORDS} words, then stratified by duration:")
    sample_rows = stratified_sample(out_rows)
    for i, row in enumerate(sample_rows, start=1):
        row = dict(row)
        row["sr_no"] = i
        sample_rows[i - 1] = row
    with open(SAMPLE_CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(sample_rows)
    write_xlsx(fieldnames, sample_rows, SAMPLE_XLSX_PATH)

    print(f"[review] threshold: WER > {WER_THRESHOLD:.0f}% on {MODE}, required models: "
          f"{', '.join(REQUIRED_MODELS)} (+{BONUS_MODEL} as a bonus signal, not required)")
    print("[review] per-model flagged clip counts (WER > threshold):")
    for m in ALL_MODELS:
        req = "required" if m in REQUIRED_MODELS else "bonus"
        print(f"    {m:14s} {flagged_counts[m]:4d}  ({req})")
    print(f"[review] wrote {len(out_rows)} rows to {CSV_PATH}")
    print(f"[review] wrote reviewer sheet (with dropdowns) to {XLSX_PATH}")
    print(f"[sample] wrote {len(sample_rows)} rows to {SAMPLE_CSV_PATH} and {SAMPLE_XLSX_PATH}")


def write_xlsx(fieldnames, rows, path: str = XLSX_PATH) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(fieldnames)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    check_fill = PatternFill("solid", fgColor="FFF2CC")
    flag_fill = PatternFill("solid", fgColor="E2EFDA")
    decision_fill = PatternFill("solid", fgColor="FCE4D6")

    check_cols = {"reference_check", "corrected_reference"} | {
        f"hyp_{m}_check" for m in ALL_MODELS
    }
    decision_cols = {"reviewer_decision", "reviewer_notes"}

    for c, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        if name in check_cols:
            cell.fill = check_fill
        elif name == "error_type":
            cell.fill = flag_fill
            # No dropdown here (see add_dropdown below): Numbers/Sheets don't
            # honor Excel's "warning" (soft) validation style the way Excel
            # does, so a data-validated cell there becomes an unescapable
            # single-pick, which defeats "type more than one, comma-separated".
            # The option vocabulary is attached as a header comment instead,
            # a plain hover note has no enforcement behavior to disagree on.
            cell.comment = Comment(
                "Free text. Pick from these, comma-separated if more than one applies:\n"
                + "\n".join(f"- {o}" for o in ERROR_TYPE_OPTIONS),
                "review sheet",
            )
        elif name in decision_cols:
            cell.fill = decision_fill
        else:
            cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append([row[k] for k in fieldnames])

    for r in range(2, len(rows) + 2):
        for c in range(1, len(fieldnames) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "sr_no": 6, "sample_id": 12, "audio_filename": 16, "reference": 45,
        "avg_wer": 9, "n_models_flagged": 9, "native_language": 12,
        "duration_seconds": 10, "ref_words": 9, "reference_check": 16, "corrected_reference": 40,
        "error_type": 30, "reviewer_decision": 20, "reviewer_notes": 30,
    }
    for c, name in enumerate(fieldnames, start=1):
        if name in widths:
            width = widths[name]
        elif name.startswith("hyp_") and name.endswith("_check"):
            width = 14
        elif name.startswith("hyp_"):
            width = 35
        elif name.startswith("wer_"):
            width = 8
        else:
            width = 12
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "E2"  # keep sr_no/sample_id/audio_filename/reference visible while scrolling

    def add_dropdown(colname, options):
        c = fieldnames.index(colname) + 1
        letter = get_column_letter(c)
        formula = '"' + ",".join(options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.error = "Pick one of the listed options."
        dv.errorTitle = "Invalid entry"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")

    add_dropdown("reference_check", REFERENCE_CHECK_OPTIONS)
    for m in ALL_MODELS:
        add_dropdown(f"hyp_{m}_check", HYP_CHECK_OPTIONS)
    # error_type deliberately has no dropdown, see the comment attached to its
    # header cell above.
    add_dropdown("reviewer_decision", REVIEWER_DECISION_OPTIONS)

    wb.save(path)


if __name__ == "__main__":
    main()
