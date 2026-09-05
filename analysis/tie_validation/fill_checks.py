"""Fill the review sheet's auto-derivable columns from the reviewer's corrected_reference.

Takes a review CSV where corrected_reference has already been hand-filled (the
reviewer listened to each clip and typed the true transcript) and:

  1. Adds normalised_corrected_reference: corrected_reference run through a
     math-notation-to-words pass (Greek letters, sub/superscripts, operators,
     see convert_math_notation) followed by the project's standard
     transcript_clean normalizer (utils.normalize.normalize_text), so it is
     directly comparable to how model hypotheses are normalized elsewhere in
     this repo under the transcript_clean mode.
  2. Fills reference_check by comparing the (same-normalized) `reference`
     column against normalised_corrected_reference.
  3. Fills each hyp_<model>_check by comparing the (same-normalized) hypothesis
     against normalised_corrected_reference.
  4. Fills error_type, reviewer_decision, and reviewer_notes, but only with
     what a text diff can actually support without hearing the clip:
       - reference_word_recall(normalised_corrected_reference, reference)
         below MISALIGNMENT_RECALL_MAX flags the reference as describing
         different content entirely (not just a noisy transcription of the
         same clip) => error_type "Misalignment", decision "Reference error".
       - reference wrong but most hyps agree with the truth => "Reference
         error"; reference roughly right but most hyps wrong => "Genuine
         model error". Otherwise reviewer_decision is left blank.
       - a multi-digit number or leftover math/Greek notation in the
         corrected transcript => "Technical vocabulary" (also noted, since
         long IDs spoken digit-by-digit vs. spelled out by the normalizer as
         one cardinal can inflate WER without a real hypothesis error).
       - an adjacent repeated word run in the corrected transcript =>
         "Disfluency" (it survived manual correction, so it's genuine
         speech, not a mis-transcription).
     Noise / speed / accent / code-switching and any "Audio artifact" /
     "Not a real error" / "Unsure" decision are NOT guessed here, they need
     the audio; those cells are left for the human reviewer.

Classification is WER-threshold based (<=8% Correct, <=40% Partially correct,
>40% Incorrect) then hand-checked row by row against a printed diff report
before being accepted (see review_report.txt in the same directory).

Usage:
    uv run --with jiwer --with num2words --with openpyxl python3 \\
        analysis/tie_validation/fill_checks.py --in <path to filled csv>
"""

import argparse
import csv
import os
import re
import string
import sys
import unicodedata

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from utils.normalize import normalize_text
from utils.wer_compute import reference_word_recall

HERE = os.path.dirname(__file__)
ALL_MODELS = ["large", "parakeet", "parakeet_ctc", "qwen3", "medium"]

CORRECT_MAX_WER = 0.08
PARTIAL_MAX_WER = 0.40

# Below this, the true transcript and the dataset reference share so few words
# that the reference looks like it describes different content entirely, not
# just a noisy/garbled transcription of the same clip. Picked from the actual
# gap in this sheet's distribution: genuine topic-mismatch rows sit at
# 0.10-0.32 recall, everything else is 0.42+.
MISALIGNMENT_RECALL_MAX = 0.35

LONG_NUMBER_RE = re.compile(r"\b\d{4,}\b")

SUBSCRIPT_MAP = {
    "₀": "0", "₁": "1", "₂": "2", "₃": "3", "₄": "4",
    "₅": "5", "₆": "6", "₇": "7", "₈": "8", "₉": "9",
    "ᵢ": "i", "ⱼ": "j", "ₖ": "k", "ₙ": "n",
    "ᵣ": "r", "ₚ": "p", "ᵥ": "v",
}
GREEK_MAP = {
    "π": "pi", "ρ": "rho", "θ": "theta", "ξ": "xi",
    "μ": "mu", "ε": "epsilon", "β": "beta", "Γ": "gamma",
    "ω": "omega",
}
MATH_NOTATION_CHARS = (
    set(SUBSCRIPT_MAP) | set(GREEK_MAP) | set("²³¹ⁿ⁺⁻′×−≠_") | {"̂", "̃"}
)


def find_long_numbers(text: str) -> list[str]:
    """IDs/codes spoken digit-by-digit (e.g. an IC part number "74138"), as
    opposed to a genuine quantity. The normalizer spells any number out as a
    single cardinal ("seventy-four thousand..."), which won't match a model
    that (correctly) transcribed it digit by digit, and can inflate WER here
    without that being a real hypothesis error.
    """
    return LONG_NUMBER_RE.findall(text)


def find_repeated_phrase(text: str) -> str | None:
    """An adjacent repeated 2-4 word run in the human-corrected transcript,
    a text signature of genuine spoken disfluency (stammer/restart) rather
    than a transcription mistake, since it survived manual correction.
    """
    words = [w.strip(string.punctuation).lower() for w in text.split()]
    words = [w for w in words if w]
    for n in (4, 3, 2):
        for i in range(len(words) - 2 * n + 1):
            if words[i:i + n] == words[i + n:i + 2 * n]:
                return " ".join(words[i:i + n])
    return None


def convert_math_notation(text: str) -> str:
    """Rewrite Greek letters, sub/superscripts, and math operators as spoken
    English, so a WER comparison against ASR hypotheses (which only ever
    output spoken words) isn't penalized for symbolic notation alone.

    Mapping is grounded in the exact symbol set actually used in this sheet's
    corrected_reference column (checked once via a full character scan), not
    a generic math-to-text library.
    """
    if not text:
        return text
    text = unicodedata.normalize("NFC", text)

    text = text.replace("_", " ")  # subscript join, e.g. I_bias -> I bias

    text = re.sub(r"(\w)̂", r"\1 hat ", text)   # combining circumflex, beta-hat
    text = re.sub(r"(\w)̃", r"\1 tilde ", text)  # combining tilde, u-tilde
    text = text.replace("ẍ", " x double dot ")   # composed x-with-diaeresis (physics ddot notation)

    text = text.replace("⁻¹", " inverse ")  # superscript minus-one, A^-1

    text = text.replace("⁺", " plus ")   # superscript plus
    text = text.replace("⁻", " minus ")  # superscript minus

    text = text.replace("²", " squared ")
    text = text.replace("³", " cubed ")
    text = text.replace("¹", " to the power one ")
    text = text.replace("ⁿ", " to the power n ")

    for ch, word in SUBSCRIPT_MAP.items():
        text = text.replace(ch, " " + word + " ")

    text = text.replace("′", " prime")  # derivative prime, u'

    for ch, word in GREEK_MAP.items():
        text = text.replace(ch, " " + word + " ")

    text = text.replace("×", " into ")    # multiplication sign
    text = text.replace("−", " minus ")   # unicode minus sign
    text = re.sub(r"(?<=\s)-(?=\s)", " minus ", text)  # ascii hyphen used as subtraction (spaced)
    text = text.replace("=", " equal to ")
    text = text.replace("/", " by ")
    text = text.replace("%", " percent ")
    text = text.replace("+", " plus ")
    text = text.replace("*", " star ")
    text = text.replace("≠", " not equal to ")

    return text


def normalize_for_compare(text: str) -> str:
    return normalize_text(convert_math_notation(text or ""))


def word_wer(ref: str, hyp: str):
    import jiwer
    if not ref and not hyp:
        return 0.0
    if not ref:
        return 1.0 if hyp else 0.0
    if not hyp:
        return 1.0
    return jiwer.wer(ref, hyp)


def classify(wer: float) -> str:
    if wer <= CORRECT_MAX_WER:
        return "Correct"
    if wer <= PARTIAL_MAX_WER:
        return "Partially correct"
    return "Incorrect"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_path", required=True)
    ap.add_argument("--out-csv", default=os.path.join(HERE, "review_sheet.csv"))
    ap.add_argument("--out-xlsx", default=os.path.join(HERE, "review_sheet.xlsx"))
    ap.add_argument("--report", default=os.path.join(HERE, "review_report.txt"))
    args = ap.parse_args()

    with open(args.in_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    idx = fieldnames.index("corrected_reference")
    fieldnames = fieldnames[:idx + 1] + ["normalised_corrected_reference"] + fieldnames[idx + 1:]

    report_lines = []

    for row in rows:
        ncr_raw = row["corrected_reference"]
        ncr = normalize_for_compare(ncr_raw)
        row["normalised_corrected_reference"] = ncr

        report_lines.append(f"=== row {row['sr_no']}  ({row['sample_id']}) ===")
        report_lines.append(f"normalised_corrected_reference: {ncr}")

        ref_norm = normalize_for_compare(row["reference"])
        ref_wer = word_wer(ncr, ref_norm)
        row["reference_check"] = classify(ref_wer)
        report_lines.append(f"reference_check   = {row['reference_check']:18s} (wer={ref_wer:.2f})")
        report_lines.append(f"  reference (norm): {ref_norm}")

        hyp_checks = {}
        for m in ALL_MODELS:
            hyp_norm = normalize_for_compare(row[f"hyp_{m}"])
            hyp_wer = word_wer(ncr, hyp_norm)
            check = classify(hyp_wer)
            hyp_checks[m] = check
            row[f"hyp_{m}_check"] = check
            report_lines.append(f"hyp_{m:13s}_check = {check:18s} (wer={hyp_wer:.2f})")
            report_lines.append(f"  hyp_{m} (norm): {hyp_norm}")

        n_models_ok = sum(1 for c in hyp_checks.values() if c in ("Correct", "Partially correct"))
        n_models_incorrect = sum(1 for c in hyp_checks.values() if c == "Incorrect")

        # How much of the true transcript's own vocabulary shows up anywhere
        # in the dataset reference: low => reference describes different
        # content, not just a noisy version of the same clip.
        ref_recall = reference_word_recall(ncr, ref_norm)
        is_misaligned = ref_recall < MISALIGNMENT_RECALL_MAX

        error_types = []
        notes = []

        if is_misaligned:
            error_types.append("Misalignment (wrong clip boundary)")
            notes.append(
                f"Reference shares only {ref_recall:.0%} of the true transcript's words, "
                "looks like it describes different content, not just a noisy transcription."
            )
        elif row["reference_check"] != "Correct":
            error_types.append("Reference error")

        long_nums = find_long_numbers(ncr_raw)
        if long_nums:
            error_types.append("Technical vocabulary (jargon, numbers, names)")
            notes.append(
                f"Contains a multi-digit number ({', '.join(long_nums)}) likely spoken "
                "digit by digit; the normalizer spells it as one large number, which can "
                "inflate WER here without a real hypothesis error."
            )
        elif any(ch in MATH_NOTATION_CHARS for ch in ncr_raw):
            error_types.append("Technical vocabulary (jargon, numbers, names)")

        repeat = find_repeated_phrase(ncr_raw)
        if repeat:
            error_types.append("Disfluency (fillers, repetitions, false starts)")
            notes.append(
                f"True transcript repeats the phrase \"{repeat}\", looks like genuine "
                "disfluency, not a transcription error."
            )

        row["error_type"] = ", ".join(dict.fromkeys(error_types))
        row["reviewer_notes"] = " ".join(notes)

        decision = ""
        if is_misaligned:
            decision = "Reference error"
        elif row["reference_check"] != "Correct" and n_models_ok >= 3:
            decision = "Reference error"
        elif row["reference_check"] in ("Correct", "Partially correct") and n_models_incorrect >= 3:
            decision = "Genuine model error"
        row["reviewer_decision"] = decision

        report_lines.append(f"ref_recall = {ref_recall:.2f}   error_type -> {row['error_type'] or '(none)'}")
        report_lines.append(f"reviewer_decision -> {decision or '(left blank, ambiguous)'}")
        if row["reviewer_notes"]:
            report_lines.append(f"reviewer_notes -> {row['reviewer_notes']}")
        report_lines.append("")

    with open(args.out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"[fill] wrote {len(rows)} rows to {args.out_csv}")

    with open(args.report, "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))
    print(f"[fill] wrote diff report to {args.report}")

    write_xlsx(fieldnames, rows, args.out_xlsx)
    print(f"[fill] wrote reviewer sheet (with dropdowns) to {args.out_xlsx}")


def write_xlsx(fieldnames, rows, xlsx_path) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    REFERENCE_CHECK_OPTIONS = ["Correct", "Partially correct", "Incorrect"]
    HYP_CHECK_OPTIONS = ["Correct", "Partially correct", "Incorrect"]
    REVIEWER_DECISION_OPTIONS = [
        "Genuine model error", "Reference error", "Audio artifact",
        "Not a real error", "Unsure",
    ]
    ERROR_TYPE_OPTIONS = [
        "Noise / audio quality", "Speed (fast or slow / unclear)",
        "Accent / pronunciation", "Technical vocabulary (jargon, numbers, names)",
        "Disfluency (fillers, repetitions, false starts)",
        "Code-switching (non-English words)", "Reference error",
        "Misalignment (wrong clip boundary)", "Other",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(fieldnames)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    check_fill = PatternFill("solid", fgColor="FFF2CC")
    flag_fill = PatternFill("solid", fgColor="E2EFDA")
    decision_fill = PatternFill("solid", fgColor="FCE4D6")

    check_cols = {"reference_check", "corrected_reference", "normalised_corrected_reference"} | {
        f"hyp_{m}_check" for m in ALL_MODELS
    }
    decision_cols = {"reviewer_decision", "reviewer_notes"}

    for c, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        if name in check_cols:
            cell.fill = check_fill
        elif name == "error_type":
            cell.fill = flag_fill
            cell.comment = Comment(
                "Free text. Pick from these, comma-separated if more than one applies:\n"
                + "\n".join(f"- {o}" for o in ERROR_TYPE_OPTIONS),
                "review sheet",
            )
        elif name == "reviewer_decision":
            cell.fill = decision_fill
            # No strict dropdown here (see add_dropdown below): this column now
            # sometimes names the specific model at fault (e.g. "Reference error -
            # Large repeats..."), which a closed list can't hold. Same fix as
            # error_type: a header comment instead of enforced validation.
            cell.comment = Comment(
                "Free text. Lead with one of these, add a model name after a dash "
                "if one model in particular is at fault:\n"
                + "\n".join(f"- {o}" for o in REVIEWER_DECISION_OPTIONS),
                "review sheet",
            )
        elif name in decision_cols:
            cell.fill = decision_fill
        else:
            cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append([row[k] for k in fieldnames])

    for r in range(2, len(rows) + 2):
        for c in range(1, len(fieldnames) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "sr_no": 6, "sample_id": 12, "audio_filename": 16, "reference": 45,
        "avg_wer": 9, "avg_wer_true": 9, "n_models_flagged": 9, "native_region": 12,
        "duration_seconds": 10, "reference_check": 16, "corrected_reference": 40,
        "normalised_corrected_reference": 40,
        "error_type": 30, "reviewer_decision": 20, "reviewer_notes": 30,
    }
    for c, name in enumerate(fieldnames, start=1):
        if name in widths:
            width = widths[name]
        elif name.startswith("hyp_") and name.endswith("_check"):
            width = 14
        elif name.startswith("hyp_"):
            width = 35
        elif name.startswith("wer_"):
            width = 8
        else:
            width = 12
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "E2"

    def add_dropdown(colname, options):
        c = fieldnames.index(colname) + 1
        letter = get_column_letter(c)
        formula = '"' + ",".join(options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.error = "Pick one of the listed options."
        dv.errorTitle = "Invalid entry"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")

    add_dropdown("reference_check", REFERENCE_CHECK_OPTIONS)
    for m in ALL_MODELS:
        add_dropdown(f"hyp_{m}_check", HYP_CHECK_OPTIONS)
    # reviewer_decision deliberately has no dropdown, see the comment attached to
    # its header cell above.

    wb.save(xlsx_path)


if __name__ == "__main__":
    main()
