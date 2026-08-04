"""Build a non-blind human review sheet of TIE_shorts clips that are hard for
several strong models at once.

Shows the annotator every model's hypothesis so they can judge, per clip,
whether the high WER comes from the audio, the reference transcript, or the
models themselves. Clips are selected by requiring several of the strongest
available models to agree a clip is hard, which filters out model-specific
failure modes.

Review columns, all reviewer-facing (see write_xlsx for the exact layout):
    reference_check     is the reference transcript itself correct
    corrected_reference  free text, fill in only if reference_check != Correct
    hyp_<model>_check   per-model correctness (a clip is rarely uniformly
                         right or wrong across all 5 models, so this is
                         tracked per model rather than one aggregate column)
    error_type          one cell, but a clip can have more than one cause:
                         pick from the dropdown for a single cause, or type
                         two or more options separated by commas
    reviewer_decision   final one-word verdict for the row
    reviewer_notes      free text, optional

Writes two files:
    review_sheet.csv   source of truth, plain CSV
    review_sheet.xlsx  same data, reviewer-facing, with the review columns
                        constrained to dropdown lists (data validation)

Usage:
    uv run --with openpyxl python3 analysis/tie_validation/build_sample.py
"""

import os
import sys
import csv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from utils.io_helpers import stage2_dir

MODE = "transcript_clean"
REQUIRED_MODELS = ["large", "parakeet", "parakeet_ctc", "qwen3"]
BONUS_MODEL = "medium"
ALL_MODELS = REQUIRED_MODELS + [BONUS_MODEL]
WER_THRESHOLD = 40.0  # percent
MIN_MODELS_FLAGGED = 3  # of len(REQUIRED_MODELS)

HERE = os.path.dirname(__file__)
CSV_PATH = os.path.join(HERE, "review_sheet.csv")
XLSX_PATH = os.path.join(HERE, "review_sheet.xlsx")

# Dropdown options. Kept short and closed-set on purpose: a fixed vocabulary
# is what makes the filled sheet directly tabulable afterward, instead of
# free text that needs re-normalizing before it can be analyzed.
REFERENCE_CHECK_OPTIONS = ["Correct", "Partially correct", "Incorrect"]
HYP_CHECK_OPTIONS = ["Correct", "Partially correct", "Incorrect"]
REVIEWER_DECISION_OPTIONS = [
    "Genuine model error",
    "Reference error",
    "Audio artifact",
    "Not a real error",
    "Unsure",
]

# error_type is one column but supports more than one cause per clip: the
# dropdown offers these as single picks, and typing two or more separated by
# commas (e.g. "Noise, Technical vocabulary") is accepted, just flagged with
# a soft warning instead of being blocked, since Excel/xlsx data validation
# has no native multi-select-into-one-cell control.
ERROR_TYPE_OPTIONS = [
    "Noise / audio quality",
    "Speed (fast or slow / unclear)",
    "Accent / pronunciation",
    "Technical vocabulary (jargon, numbers, names)",
    "Disfluency (fillers, repetitions, false starts)",
    "Code-switching (non-English words)",
    "Reference error",
    "Misalignment (wrong clip boundary)",
    "Other",
]


def load_model(model: str) -> dict:
    """Return {sample_id: row_dict} from that model's WER csv."""
    path = os.path.join(stage2_dir("tie"), MODE, f"wer_{model}_{MODE}.csv")
    rows = {}
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows[str(row["ID"])] = row
    return rows


def main() -> None:
    tables = {m: load_model(m) for m in ALL_MODELS}
    base_ids = list(tables[REQUIRED_MODELS[0]].keys())

    flagged_counts = {m: 0 for m in ALL_MODELS}
    out_rows = []

    for sid in base_ids:
        base_row = tables[REQUIRED_MODELS[0]][sid]
        wers = {}
        hyps = {}
        for m in ALL_MODELS:
            r = tables[m].get(sid)
            if r is None:
                continue
            w = float(r["wer"]) * 100
            wers[m] = w
            hyps[m] = r["hypothesis_raw"]
            if w > WER_THRESHOLD:
                flagged_counts[m] += 1

        n_models_flagged = sum(
            1 for m in REQUIRED_MODELS if m in wers and wers[m] > WER_THRESHOLD
        )
        if n_models_flagged < MIN_MODELS_FLAGGED:
            continue

        avg_wer = round(sum(wers[m] for m in wers) / len(wers), 2) if wers else ""

        row = {
            "sample_id": sid,
            "audio_filename": f"{sid}.wav",
            "reference": base_row["reference_raw"],
            **{f"hyp_{m}": hyps.get(m, "") for m in ALL_MODELS},
            **{f"wer_{m}": round(wers[m], 2) if m in wers else "" for m in ALL_MODELS},
            "avg_wer": avg_wer,
            "n_models_flagged": n_models_flagged,
            "native_region": base_row.get("Native_Region", ""),
            "duration_seconds": base_row.get("Speech_Duration_seconds", ""),
            "reference_check": "",
            "corrected_reference": "",
            **{f"hyp_{m}_check": "" for m in ALL_MODELS},
            "error_type": "",
            "reviewer_decision": "",
            "reviewer_notes": "",
        }
        out_rows.append(row)

    out_rows.sort(key=lambda r: r["n_models_flagged"], reverse=True)
    for i, row in enumerate(out_rows, start=1):
        row["sr_no"] = i

    fieldnames = (
        ["sr_no", "sample_id", "audio_filename", "reference"]
        + [f"hyp_{m}" for m in ALL_MODELS]
        + [f"wer_{m}" for m in ALL_MODELS]
        + ["avg_wer", "n_models_flagged", "native_region", "duration_seconds",
           "reference_check", "corrected_reference"]
        + [f"hyp_{m}_check" for m in ALL_MODELS]
        + ["error_type", "reviewer_decision", "reviewer_notes"]
    )

    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    write_xlsx(fieldnames, out_rows)

    print(f"[review] threshold: WER > {WER_THRESHOLD:.0f}% on {MODE}, required models: "
          f"{', '.join(REQUIRED_MODELS)} (+{BONUS_MODEL} as a bonus signal, not required)")
    print("[review] per-model flagged clip counts (WER > threshold):")
    for m in ALL_MODELS:
        req = "required" if m in REQUIRED_MODELS else "bonus"
        print(f"    {m:14s} {flagged_counts[m]:4d}  ({req})")
    print(f"[review] wrote {len(out_rows)} rows to {CSV_PATH}")
    print(f"[review] wrote reviewer sheet (with dropdowns) to {XLSX_PATH}")


def write_xlsx(fieldnames, rows) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(fieldnames)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    check_fill = PatternFill("solid", fgColor="FFF2CC")
    flag_fill = PatternFill("solid", fgColor="E2EFDA")
    decision_fill = PatternFill("solid", fgColor="FCE4D6")

    check_cols = {"reference_check", "corrected_reference"} | {
        f"hyp_{m}_check" for m in ALL_MODELS
    }
    decision_cols = {"reviewer_decision", "reviewer_notes"}

    for c, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        if name in check_cols:
            cell.fill = check_fill
        elif name == "error_type":
            cell.fill = flag_fill
            # No dropdown here (see add_dropdown below): Numbers/Sheets don't
            # honor Excel's "warning" (soft) validation style the way Excel
            # does, so a data-validated cell there becomes an unescapable
            # single-pick, which defeats "type more than one, comma-separated".
            # The option vocabulary is attached as a header comment instead,
            # a plain hover note has no enforcement behavior to disagree on.
            cell.comment = Comment(
                "Free text. Pick from these, comma-separated if more than one applies:\n"
                + "\n".join(f"- {o}" for o in ERROR_TYPE_OPTIONS),
                "review sheet",
            )
        elif name in decision_cols:
            cell.fill = decision_fill
        else:
            cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        ws.append([row[k] for k in fieldnames])

    for r in range(2, len(rows) + 2):
        for c in range(1, len(fieldnames) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "sr_no": 6, "sample_id": 12, "audio_filename": 16, "reference": 45,
        "avg_wer": 9, "n_models_flagged": 9, "native_region": 12,
        "duration_seconds": 10, "reference_check": 16, "corrected_reference": 40,
        "error_type": 30, "reviewer_decision": 20, "reviewer_notes": 30,
    }
    for c, name in enumerate(fieldnames, start=1):
        if name in widths:
            width = widths[name]
        elif name.startswith("hyp_") and name.endswith("_check"):
            width = 14
        elif name.startswith("hyp_"):
            width = 35
        elif name.startswith("wer_"):
            width = 8
        else:
            width = 12
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "E2"  # keep sr_no/sample_id/audio_filename/reference visible while scrolling

    def add_dropdown(colname, options):
        c = fieldnames.index(colname) + 1
        letter = get_column_letter(c)
        formula = '"' + ",".join(options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.error = "Pick one of the listed options."
        dv.errorTitle = "Invalid entry"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")

    add_dropdown("reference_check", REFERENCE_CHECK_OPTIONS)
    for m in ALL_MODELS:
        add_dropdown(f"hyp_{m}_check", HYP_CHECK_OPTIONS)
    # error_type deliberately has no dropdown, see the comment attached to its
    # header cell above.
    add_dropdown("reviewer_decision", REVIEWER_DECISION_OPTIONS)

    wb.save(XLSX_PATH)


if __name__ == "__main__":
    main()
